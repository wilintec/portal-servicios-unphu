#!/usr/bin/env python3
"""Genera index.html desde la versión más reciente del Excel maestro Servicios*.xlsx.

Uso local:
    python scripts/generar_portal.py
    python scripts/generar_portal.py --excel datos/Servicios.xlsx

El script abre el Excel solamente en modo lectura. Conserva el diseño de index.html
y reemplaza únicamente el bloque delimitado por BEGIN_GENERATED_DATA y
END_GENERATED_DATA.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_DATA_DIR = Path("datos")
DEFAULT_HTML = Path("index.html")
SHEET_NAME = "servicios"
MONTHS_ES = (
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
)

CANONICAL_HEADERS = {
    "laboratorio": "Laboratorio",
    "contacto": "Contacto",
    "servicio": "Servicio",
    "requiere equipo del Laboratorio": "Requiere Equipo del Laboratorio",
    "descripcion y alcance": "Descripcion y Alcance",
    "tipo de usuario al que podria dirigirse": "Tipo de usuario",
    "equipos requeridos para su prestacion": "Equipos requeridos",
    "materiales reactivos o insumos necesarios": "Materiales",
    "personal tecnico o especializado requerido": "Personal",
    "tiempo estimado para la realizacion del servicio": "Tiempo estimado",
    "cualquier requisito limitacion o condicion especial que deba considerarse para el arrendamiento del servicio": "Condiciones",
}


def clean(value: Any) -> str:
    """Convierte una celda a texto limpio y transforma saltos de línea en espacios."""
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())


def normalized(value: Any) -> str:
    """Normaliza mayúsculas, acentos, signos y espacios para comparar encabezados."""
    text = unicodedata.normalize("NFKD", clean(value).casefold())
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def natural_key(path: Path) -> tuple[Any, ...]:
    """Orden natural: ServiciosV10.xlsx queda después de ServiciosV9.xlsx."""
    parts = re.split(r"(\d+)", path.stem.casefold())
    return tuple(int(part) if part.isdigit() else part for part in parts)


def discover_excel(data_dir: Path) -> Path:
    candidates = [
        path for path in data_dir.glob("Servicios*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No se encontró ningún archivo Servicios*.xlsx en {data_dir}."
        )
    return sorted(candidates, key=natural_key)[-1]


def find_sheet(workbook: Any, expected_name: str) -> Any:
    for name in workbook.sheetnames:
        if normalized(name) == normalized(expected_name):
            return workbook[name]
    raise KeyError(
        f"No existe la hoja {expected_name!r}. Hojas disponibles: {workbook.sheetnames}"
    )


def publication_allowed(value: Any) -> bool:
    return normalized(value) in {"si", "s", "true", "1", "x", "publicar"}


def parse_requires_equipment(value: Any) -> bool:
    return normalized(value) in {"si", "s", "true", "1", "x", "requiere", "requerido"}


def read_services(excel_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(
        excel_path,
        data_only=True,
        read_only=True,
        keep_links=False,
    )
    worksheet = find_sheet(workbook, SHEET_NAME)
    rows = worksheet.iter_rows(values_only=True)

    try:
        raw_headers = list(next(rows))
    except StopIteration as exc:
        raise ValueError("El Excel está vacío.") from exc

    normalized_headers: dict[str, int] = {}
    for index, header in enumerate(raw_headers):
        key = normalized(header)
        if key:
            normalized_headers[key] = index

    missing = [
        display_name
        for normalized_name, display_name in CANONICAL_HEADERS.items()
        if normalized_name not in normalized_headers
    ]
    if missing:
        found = ", ".join(clean(value) for value in raw_headers if clean(value))
        raise KeyError(
            "Faltan columnas obligatorias: " + ", ".join(missing) +
            f". Encabezados encontrados: {found}"
        )

    publish_index = normalized_headers.get("publicar")

    def value(values: tuple[Any, ...], normalized_header: str) -> Any:
        index = normalized_headers[normalized_header]
        return values[index] if index < len(values) else None

    services: list[dict[str, Any]] = []
    for excel_row_number, values in enumerate(rows, start=2):
        if publish_index is not None:
            publish_value = values[publish_index] if publish_index < len(values) else None
            if not publication_allowed(publish_value):
                continue

        laboratory = clean(value(values, "laboratorio"))
        service_type = clean(value(values, "servicio"))
        description = clean(value(values, "descripcion y alcance"))

        if not any((laboratory, service_type, description)):
            continue
        if not laboratory:
            raise ValueError(f"Fila {excel_row_number}: falta el laboratorio.")
        if not service_type:
            raise ValueError(f"Fila {excel_row_number}: falta el tipo de servicio.")
        if not description:
            raise ValueError(f"Fila {excel_row_number}: falta la descripción y alcance.")

        services.append({
            "id": len(services) + 1,
            "sourceRow": excel_row_number,
            "laboratory": laboratory,
            "contact": clean(value(values, "contacto")),
            "serviceType": service_type,
            "requiresEquipment": parse_requires_equipment(
                value(values, "requiere equipo del Laboratorio")
            ),
            "description": description,
            "audience": clean(value(values, "tipo de usuario al que podria dirigirse")),
            "equipment": clean(value(values, "equipos requeridos para su prestacion")),
            "materials": clean(value(values, "materiales reactivos o insumos necesarios")),
            "personnel": clean(value(values, "personal tecnico o especializado requerido")),
            "duration": clean(value(values, "tiempo estimado para la realizacion del servicio")),
            "conditions": clean(value(values, "cualquier requisito limitacion o condicion especial que deba considerarse para el arrendamiento del servicio")),
        })

    if not services:
        publication_note = (
            " con Publicar = Sí" if publish_index is not None else ""
        )
        raise ValueError(f"No se encontraron filas válidas{publication_note}.")

    services.sort(
        key=lambda item: (
            item["laboratory"].casefold(),
            item["serviceType"].casefold(),
            item["sourceRow"],
        )
    )
    for item_id, item in enumerate(services, start=1):
        item["id"] = item_id
    return services


def format_date_es(value: date) -> str:
    return f"{value.day} de {MONTHS_ES[value.month - 1]} de {value.year}"


def update_html(
    html_path: Path,
    services: list[dict[str, Any]],
    source_file: str,
) -> None:
    html = html_path.read_text(encoding="utf-8")
    metadata = {
        "updatedAt": format_date_es(date.today()),
        "sourceFile": source_file,
    }
    generated_block = (
        "/* BEGIN_GENERATED_DATA */\n"
        f"const services={json.dumps(services, ensure_ascii=False, indent=2)};\n"
        f"const portalMeta={json.dumps(metadata, ensure_ascii=False, separators=(',', ':'))};\n"
        "/* END_GENERATED_DATA */"
    )

    pattern = re.compile(
        r"/\* BEGIN_GENERATED_DATA \*/.*?/\* END_GENERATED_DATA \*/",
        flags=re.DOTALL,
    )
    html, replacement_count = pattern.subn(
        lambda _: generated_block,
        html,
        count=1,
    )
    if replacement_count != 1:
        raise RuntimeError(
            "No fue posible localizar el bloque de datos generado en index.html. "
            "No elimines los marcadores BEGIN_GENERATED_DATA y END_GENERATED_DATA."
        )

    html_path.write_text(html, encoding="utf-8", newline="\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--excel",
        type=Path,
        default=None,
        help="Excel concreto. Si se omite, se usa la versión Servicios*.xlsx más alta.",
    )
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--html", type=Path, default=DEFAULT_HTML)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        excel_path = args.excel or discover_excel(args.data_dir)
        services = read_services(excel_path)
        update_html(args.html, services, excel_path.name)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    laboratories = len({item["laboratory"] for item in services})
    service_types = len({item["serviceType"] for item in services})
    equipment_services = sum(item["requiresEquipment"] for item in services)
    print(
        "Portal actualizado: "
        f"{args.html} | fuente={excel_path} | servicios={len(services)} | "
        f"laboratorios={laboratories} | tipos={service_types} | "
        f"requieren_equipos={equipment_services}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
